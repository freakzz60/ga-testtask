import main
import openpyxl as oxl
import pandas as pd


def maketable(filename, firsttable, secondtable):
    result = firsttable.join(secondtable, how='right', lsuffix=' USD/RUB', rsuffix=' JPY/RUB')
    # Объединить две таблицы
    try:
        result.to_excel(filename, index=False)
    except Exception:
        print('Не получилось записать таблицу на диск. Нет прав/закончилось место?')
    # Записать файл без индексов
    calcresult(filename)
    # Поделить USD/RUB на JPY/RUB
    formatting(filename)
    # Отформатировать таблицу в соответствии с заданием


def calcresult(filename):
    df = pd.read_excel(filename)
    df['Результат'] = df['Курс USD/RUB'] / df['Курс JPY/RUB']
    df.to_excel(filename, index=False)


def formatting(filename):
    wb = oxl.load_workbook(filename=filename)
    ws = wb.active
    column_letters = tuple(oxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
    for column_letter in column_letters:
        ws.column_dimensions[column_letter].auto_size = True
        ws.column_dimensions[column_letter].width += 2
    for row in (ws.iter_rows()):
        for c in row:
            c.number_format = '₽ #,##0.00'
    wb.save(filename)
    wb.close()


def countrows(filename):
    wb = oxl.load_workbook(filename=filename)
    ws = wb.active
    rowsnum = str(ws.max_row - 1)
    return rowsnum
